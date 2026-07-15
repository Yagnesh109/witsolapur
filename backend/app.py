import os
import uuid
import datetime
from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS

app = Flask(__name__)
# Enable CORS for all routes so that the static frontend can connect from file:// or other local servers
CORS(app)

# Configuration
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'alumni_images')
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
ADMIN_PASSCODE = os.environ.get('ADMIN_PASSCODE', 'admin123')

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Helper to check image extension
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# --- Database Manager for JSON Database ---
class DatabaseManager:
    def __init__(self, db_path='alumni_db.json'):
        import json
        # Resolve path relative to backend directory
        base_dir = os.path.dirname(os.path.abspath(__file__))
        self.db_path = os.path.join(base_dir, db_path)
        
        # Ensure database file exists
        if not os.path.exists(self.db_path):
            with open(self.db_path, 'w', encoding='utf-8') as f:
                json.dump([], f)
        print(f"[INFO] JSON database is active at: {self.db_path}")

    def get_all(self):
        import json
        try:
            with open(self.db_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"[ERROR] Reading JSON file: {e}")
            return []

    def add(self, data):
        import json
        try:
            records = self.get_all()
            records.append(data)
            with open(self.db_path, 'w', encoding='utf-8') as f:
                json.dump(records, f, indent=2)
            return True
        except Exception as e:
            print(f"[ERROR] Writing to JSON file: {e}")
            return False

    def update_status(self, record_id, status):
        import json
        try:
            records = self.get_all()
            updated = False
            for r in records:
                if str(r.get('id')) == str(record_id):
                    r['status'] = status
                    updated = True
                    break
            if updated:
                with open(self.db_path, 'w', encoding='utf-8') as f:
                    json.dump(records, f, indent=2)
                return True
            return False
        except Exception as e:
            print(f"[ERROR] Updating JSON file: {e}")
            return False

    def delete(self, record_id):
        import json
        image_to_delete = None
        records = self.get_all()
        for r in records:
            if str(r.get('id')) == str(record_id):
                url = r.get('image_url', '')
                if url:
                    image_to_delete = url.split('/')[-1]
                break

        try:
            new_records = [r for r in records if str(r.get('id')) != str(record_id)]
            with open(self.db_path, 'w', encoding='utf-8') as f:
                json.dump(new_records, f, indent=2)
            self._delete_image_file(image_to_delete)
            return True
        except Exception as e:
            print(f"[ERROR] Deleting from JSON file: {e}")
            return False

    def _delete_image_file(self, filename):
        if filename:
            path = os.path.join(UPLOAD_FOLDER, filename)
            if os.path.exists(path):
                try:
                    os.remove(path)
                    print(f"[INFO] Deleted image from disk: {path}")
                except Exception as e:
                    print(f"[WARN] Error deleting file {path}: {e}")

db = DatabaseManager()

# --- Decorator for Admin Verification ---
def require_admin(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        passcode = request.headers.get('X-Admin-Passcode')
        if not passcode or passcode != ADMIN_PASSCODE:
            return jsonify({'success': False, 'message': 'Unauthorized admin passcode.'}), 401
        return f(*args, **kwargs)
    return decorated

# --- Routes ---

# Public: Serve uploaded images static files
@app.route('/alumni_images/<filename>')
def serve_image(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

# Public: Register Alumni
@app.route('/api/alumni/register', methods=['POST'])
def register_alumni():
    # Retrieve form data
    name = request.form.get('name')
    email = request.form.get('email')
    mobile_number = request.form.get('mobile_number')
    department = request.form.get('department')
    graduation_year = request.form.get('graduation_year')
    designation = request.form.get('designation')
    company = request.form.get('company')
    image_file = request.files.get('image')

    # Basic Validation
    if not all([name, email, mobile_number, department, graduation_year, designation, company, image_file]):
        return jsonify({'success': False, 'message': 'All fields (including mobile number and image) are required.'}), 400

    if not allowed_file(image_file.filename):
        return jsonify({'success': False, 'message': 'Invalid image format. Allowed formats: PNG, JPG, JPEG, GIF, WEBP.'}), 400

    # Save local image
    try:
        file_ext = image_file.filename.rsplit('.', 1)[1].lower()
        unique_filename = f"{uuid.uuid4().hex}.{file_ext}"
        image_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
        image_file.save(image_path)
    except Exception as e:
        return jsonify({'success': False, 'message': f'Failed to save uploaded image: {str(e)}'}), 500

    # Construct server image URL (Assuming backend runs locally on port 5000)
    image_url = f"{request.host_url}alumni_images/{unique_filename}"

    # Prepare data dict
    alumni_data = {
        'id': str(uuid.uuid4()),
        'name': name.strip(),
        'email': email.strip(),
        'mobile_number': mobile_number.strip(),
        'department': department.strip(),
        'graduation_year': graduation_year.strip(),
        'designation': designation.strip(),
        'image_url': image_url,
    }

    # Save to database
    success = db.add(alumni_data)
    if success:
        return jsonify({
            'success': True,
            'message': 'Registration submitted successfully! It will appear in the directory once verified by the administrator.'
        }), 201
    else:
        # Cleanup saved image on failure
        if os.path.exists(image_path):
            os.remove(image_path)
        return jsonify({'success': False, 'message': 'Failed to save registration data to database.'}), 500

# Public: Fetch Approved Alumni
@app.route('/api/alumni/approved', methods=['GET'])
def get_approved_alumni():
    all_alumni = db.get_all()
    # Filter by Approved status
    approved = [a for a in all_alumni if a.get('status') == 'Approved']
    # Sort by created_at descending (newest first)
    approved.sort(key=lambda x: x.get('created_at', ''), reverse=True)
    return jsonify(approved)

# Admin: Fetch All Alumni (Pending & Approved)
@app.route('/api/admin/alumni', methods=['GET'])
@require_admin
def get_all_alumni_admin():
    all_alumni = db.get_all()
    # Sort by created_at descending (newest first)
    all_alumni.sort(key=lambda x: x.get('created_at', ''), reverse=True)
    return jsonify(all_alumni)

# Admin: Export Approved Alumni as Excel sorted by year and department ascending
@app.route('/api/admin/alumni/export', methods=['GET'])
def export_alumni_excel():
    # Passcode verification via query parameter (args) or header
    passcode = request.args.get('passcode') or request.headers.get('X-Admin-Passcode')
    if not passcode or passcode != ADMIN_PASSCODE:
        return jsonify({'success': False, 'message': 'Unauthorized admin passcode.'}), 401

    all_alumni = db.get_all()
    approved = [a for a in all_alumni if a.get('status') == 'Approved']

    # Sort year-wise ascending, and branch-wise (department) ascending.
    approved.sort(key=lambda x: (
        x.get('graduation_year', '').strip(),
        x.get('department', '').strip().lower()
    ))

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Approved Alumni"

    headers = ["Name", "Email", "Mobile Number", "Department", "Graduation Year", "Designation", "Company", "Registration Date"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for alumnus in approved:
        ws.append([
            alumnus.get('name', ''),
            alumnus.get('email', ''),
            alumnus.get('mobile_number', ''),
            alumnus.get('department', ''),
            alumnus.get('graduation_year', ''),
            alumnus.get('designation', ''),
            alumnus.get('company', ''),
            alumnus.get('created_at', '')
        ])

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    import io
    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)

    from flask import send_file
    from datetime import datetime
    current_date = datetime.now().strftime("%d-%m-%Y")
    filename = f"WIT Alumni report ({current_date}).xlsx"
    return send_file(
        excel_stream,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )

# Admin: Approve Alumni
@app.route('/api/admin/alumni/<record_id>/approve', methods=['POST'])
@require_admin
def approve_alumni(record_id):
    success = db.update_status(record_id, 'Approved')
    if success:
        return jsonify({'success': True, 'message': 'Alumni registration approved successfully!'})
    else:
        return jsonify({'success': False, 'message': 'Failed to approve alumni registration or alumnus not found.'}), 404
# Admin: Delete/Reject Alumni
@app.route('/api/admin/alumni/<record_id>', methods=['DELETE'])
@require_admin
def delete_alumni(record_id):
    success = db.delete(record_id)
    if success:
        return jsonify({'success': True, 'message': 'Alumni record deleted successfully!'})
    else:
        return jsonify({'success': False, 'message': 'Failed to delete alumni record or alumnus not found.'}), 404

# Admin: Verify passcode (helper endpoint for login)
@app.route('/api/admin/verify-passcode', methods=['POST'])
def verify_passcode():
    data = request.json or {}
    passcode = data.get('passcode')
    if passcode == ADMIN_PASSCODE:
        return jsonify({'success': True, 'message': 'Passcode verification successful!'})
    else:
        return jsonify({'success': False, 'message': 'Invalid passcode.'}), 401

if __name__ == '__main__':
    # Run server locally on port 5000
    print(f"Starting WIT Alumni Backend Server on http://localhost:5000")
    print(f"Default Administrator Passcode: {ADMIN_PASSCODE}")
    app.run(host='0.0.0.0', port=5000, debug=True)
